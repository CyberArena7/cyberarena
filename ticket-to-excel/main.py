from datetime import datetime
from ratelimit.exception import RateLimitException
import repairdesk
from openpyxl import Workbook, load_workbook
from time import sleep
import re
import os

api = repairdesk.RepairDesk(os.environ["API_KEY"])


def sheet_to_dict_array(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    headers = rows[0]

    data = []
    for row in rows[1:]:
        row_dict = dict(zip(headers, row))
        data.append(row_dict)

    return data


def main():
    rd_workbook = load_workbook(filename="tickets.xlsx").active
    data = sheet_to_dict_array(rd_workbook)

    wb = Workbook()
    ws = wb.active
    assert ws is not None

    ws.append(
        [
            "N Ticket",
            "Fecha",
            "Dispositivo",
            "Estado",
            "Partes",
            "Precio partes",
            "Pagado",
            "Total",
        ]
    )

    last_id = int(data[0]["Ticket ID"].lstrip("T-")) + 1
    for row in data:
        id = row["Ticket ID"]

        for number in range(last_id - 1, int(id.lstrip("T-")), -1):
            ws.append(
                [
                    "T-{}".format(number),
                    "MISSING",
                    "MISSING",
                    "MISSING",
                    "MISSING",
                    "MISSING",
                ]
            )

        last_id = int(id.lstrip("T-"))

        internal_id = api._call("/tickets", params={"keyword": id})["data"]["ticketData"][0][
            "summary"
        ]["id"]
        ticket = api._call("/tickets/{}".format(internal_id), params={})["data"]
        device = ticket["devices"][0]

        name = device["name_with_device_and_manufacturer"]
        status = device["status"]["name"]

        created = datetime.fromtimestamp(ticket["summary"]["created_date"])

        if "25350610" in map(lambda a: a["id"], ticket["accesory"]):
            print(id, "no rush")
            continue

        total_parts = 0
        for part in device["parts"]:
            total_parts += float(part["price"]) * int(part["quantity"])

        print(
            id,
            created,
            name,
            status,
            row["Ticket Items"],
            total_parts,
            row["Paid"],
            row["Total"],
        )
        ws.append(
            [
                id,
                created,
                name,
                status,
                row["Ticket Items"],
                total_parts,
                row["Paid"],
                row["Total"],
            ]
        )
        wb.save("output.xlsx")


if __name__ == "__main__":
    main()
