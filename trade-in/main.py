from dataclasses import dataclass
from datetime import datetime
import openpyxl
from flask import Flask, request, render_template, make_response
from typing import Any
import io

app = Flask(__name__)


@dataclass
class Contact:
    name: str
    address: str
    city: str
    state: str
    post_code: str
    driving_license: str


@dataclass
class TradeIn:
    item_id: str
    transaction_id: str
    purchase_date: datetime
    sku: str
    name: str
    imei: str
    serial_number: str
    condition_on_purchase: str
    color: str
    size: str
    cost_price: int
    seller: Contact


def sheet_to_dict_array(sheet) -> list[dict[str, Any]]:
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)

    data = []
    for row in rows:
        row_dict = dict(zip(headers, row))
        if row_dict["Item ID"] is None:
            continue
        data.append(row_dict)

    return data


def product_full_name(trade: TradeIn) -> str:
    name = trade.name.upper()

    if trade.condition_on_purchase is not None:
        condition = trade.condition_on_purchase.upper()
        print(condition)
        if condition != "COMO NUEVO" and condition != "NORMAL":
            if condition not in name:
                name += " " + condition

    if trade.size is not None:
        size = trade.size.upper().replace(" ", "").replace(" ", "")
        if size not in name.replace(" ", "").replace(" ", ""):
            name += " " + size

    if trade.color is not None:
        color = trade.color.upper()
        if color not in name:
            name += " " + color

    return name


def row_to_trade_in(r):
    return TradeIn(
        item_id=r["Item ID"],
        transaction_id=r["Trans ID"],
        purchase_date=datetime.strptime(r["Purchase Date"], "%Y-%m-%d %H:%M:%S"),
        sku="",
        name=r["Name"],
        imei=r["IMEI"],
        serial_number=r["Serial Number"],
        condition_on_purchase=r["Condition On Purchase"],
        color=r["Color"],
        size=r["Size"],
        cost_price=r["Cost Price"],
        seller=Contact(
            name=" ".join(filter(None, [r["Seller First Name"], r["Seller Last Name"]])),
            driving_license=r["Seller Driving License"],
            address=r["Seller Address1"],
            city=r["Seller City"],
            state=r["Seller State"],
            post_code=r["Seller Post Code"],
        ),
    )


def sheet_to_trade_in(sheet):
    rows = sheet_to_dict_array(sheet)
    return list(
        map(
            row_to_trade_in,
            rows,
        )
    )


@app.route("/")
def index():
    return render_template("index.html")


# TODO: Zip file with all of the PDFs


@app.route("/digital", methods=["POST"])
def digital():
    book = openpyxl.load_workbook(request.files["export"])
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None

    ws.append(
        [
            "FECHA VENTA",
            "NOMBRE Y APELLIDOS",
            "DNI",
            "FECHA NACIMIENTO",
            "OBJETO",
            "IMEI 1 - SN",
            "IMEI 2",
            "P.COMPRA",
            "REF. CA",
        ]
    )

    trade_ins = sheet_to_trade_in(book.active)
    for trade in reversed(trade_ins):
        id = trade.serial_number
        if trade.serial_number is None or len(trade.serial_number) == 0:
            id = trade.imei
            trade.imei = ""

        ws.append(
            [
                trade.purchase_date,
                trade.seller.name,
                trade.seller.driving_license,
                "",
                product_full_name(trade),
                id,
                trade.imei,
                trade.cost_price,
                trade.transaction_id,
            ]
        )

    output = io.BytesIO()
    wb.save(output)

    res = make_response(output.getvalue())
    res.headers["Content-Disposition"] = "attachment; filename={} {}.xlsx".format(
        "Trade In", datetime.today().strftime("%d-%m")
    )
    res.headers["Content-type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return res


@app.route("/physical", methods=["POST"])
def physical():
    book = openpyxl.load_workbook(request.files["export"])
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None

    ws.append(
        [
            "N ORDEN",
            "FECHA",
            "APELLIDOS Y NOMBRE DEL VENDEDOR",
            "DNI O PASAPORTE",
            "DOMICILIO",
            "LOCALIDAD",
            "PROVINCIA O PAIS",
            "CLASE DE OBJETO",
            "DESCRIPCION",
            "PRECIO ABONADO",
            "FECHA DE VENTA",
        ]
    )

    trade_ins = sheet_to_trade_in(book.active)
    for trade in reversed(trade_ins):
        id = trade.serial_number
        if trade.serial_number is None or len(trade.serial_number) == 0:
            id = trade.imei

        ws.append(
            [
                "",
                trade.purchase_date,
                trade.seller.name,
                trade.seller.driving_license,
                trade.seller.address,
                "/".join(filter(None, [trade.seller.city, trade.seller.state])),
                "",
                "",
                product_full_name(trade) + " [{}]".format(id),
                trade.cost_price,
                "",
            ]
        )

    output = io.BytesIO()
    wb.save(output)

    res = make_response(output.getvalue())
    res.headers["Content-Disposition"] = "attachment; filename={} {}.xlsx".format(
        "Trade In", datetime.today().strftime("%d-%m")
    )
    res.headers["Content-type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return res


if __name__ == "__main__":
    app.run(debug=True)
